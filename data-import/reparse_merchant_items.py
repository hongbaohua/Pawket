# -*- coding: utf-8 -*-
# 用 Gemini 對 775 筆已匯入的真實資料重新判斷「商家 vs 品項（含單價）」，取代舊版
# parse_discounts.py 用括號規則硬拆的做法——舊規則沒有語意理解，同一種結構
# （例如「《遊戲名》禮包」）有時拆對有時拆錯。Ivy 逐筆核對時抓出好幾個真實錯誤：
#   - row371-373《戀與製作人》週年饋禮 / row185《戀與深空》鎏金七日禮：
#     商品名稱整個留在 merchant 裡沒拆開，應該是 merchant=遊戲名, items=[禮包名]
#   - row698「預購-酷比樂-LOOK UP...」：「預購」這種購買狀態詞被誤判成商家，
#     真正商家「酷比樂」被留在 note 裡
#   - row495/529「BABY MMMM娃娃-銀虎、河玟-0313_beomgyutxt代購（本金）」：
#     這是「代購」，需要標出代購人是誰，不是普通備註
#
# 這一版另外從原始 xlsx 挖出「金額」欄位的公式（不是算出來的最終數字）——Ivy 指出
# 當初xlsx其實是用公式表示品項/折扣計算的，例如：
#   row51  "mamamiya-《時光代理人》掛件盲盒、三麗鷗聯動美味時光收藏卡" 公式 =-220-130
#          （兩個品項的單價220/130，跟商家欄位列的兩個品項名稱對得上）
#   row267 "Lalaport-My Anime Square柯南快閃餅乾-基德、柯南"        公式 =-190*2
#          （兩個同價190的品項）
#   row300 "邵瀞葶代購-鐳塔守護神Q版-小卡鐳塔各一、顧江小立牌各一"    公式 =ROUNDUP(-(4.2*9+17.5*2)*4.45,0)
#          （匯率案例：兩個品項各自的外幣單價×數量，乘上匯率）
# 舊版 build_import2.py 的 safe_eval() 只把整條公式算成一個最終數字當 amount，
# 品項單價的結構就這樣被丟掉了。這支腳本把公式原始字串一起交給 Gemini 判斷，
# 對得上品項才填 unitPrice，對不上不硬猜、confidence 壓低讓 Ivy 自己看。
#
# 這支腳本只重新判斷 merchant / items / note / specialTag，**不**動
# grossAmount/discounts——那些已經由 parse_discounts.py 處理過、Ivy 也已經套用
# 到 Supabase，這裡直接沿用，避免兩支腳本互相打架。
#
# 產出：
#   reparse_corrections.json — 可自動套用的修正清單（含真實交易 id），給
#                               App.tsx 的一次性按鈕讀取
#   reparse_review.txt       — 修正前→修正後對照（含用到的公式），供 Ivy 抽查
#
# 使用方式：
#   1. cd Pawket/data-import
#   2. python reparse_merchant_items.py   （需要 Pawket/.env.local 裡的 GEMINI_API_KEY）
#   3. Ivy 看過 reparse_review.txt 覺得OK後，去 App「罐罐明細本」按
#      「套用商家品項重新解析(一次性)」（App.tsx handleApplyReparseCorrections，用完可移除）

import json, os, re, time

import openpyxl
from google import genai

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
IMPORTED_PATH = os.path.join(ROOT, '匯入_中信對帳_775筆.json')
DISCOUNT_CORRECTIONS_PATH = os.path.join(HERE, 'discount_corrections.json')
XLSX_PATH = os.path.join(os.path.dirname(ROOT), '對帳資料', '中國信託', '中信餘額_正確答案.xlsx')
OUT_PATH = os.path.join(HERE, 'reparse_corrections.json')
REVIEW_PATH = os.path.join(HERE, 'reparse_review.txt')

ENV_PATH = os.path.join(ROOT, '.env.local')


def load_api_key():
    with open(ENV_PATH, encoding='utf-8') as f:
        for line in f:
            if line.startswith('GEMINI_API_KEY='):
                return line.strip().split('=', 1)[1].strip()
    raise RuntimeError('GEMINI_API_KEY not found in .env.local')


def load_amount_formulas():
    """讀原始 xlsx，回傳 {row: formula_string} for 金額欄位是公式的列。"""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws = wb.active
    formulas = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=4).value
        if isinstance(v, str) and v.startswith('='):
            formulas[r] = v
    return formulas


with open(IMPORTED_PATH, encoding='utf-8') as f:
    imported = json.load(f)['transactions']

with open(DISCOUNT_CORRECTIONS_PATH, encoding='utf-8') as f:
    discount_corrections = {c['id']: c for c in json.load(f)}

amount_formulas = load_amount_formulas()
row_re = re.compile(r'row(\d+)')

# 組出每筆要丟給 Gemini 判斷的「乾淨基底文字」＋（如果有）原始金額公式：
# 有折扣修正的，用它已經拆好的 merchant+note（已經去掉折扣括號）；
# 沒有的，直接用原始 merchant（未拆過，可能還帶著沒被折扣規則吃到的括號，Gemini 自己判斷）。
rows = []
for tx in imported:
    dc = discount_corrections.get(tx['id'])
    if dc:
        base_text = dc['merchant'] + (f"-{dc['note']}" if dc.get('note') else '')
        current_merchant, current_note = dc['merchant'], dc.get('note')
    else:
        base_text = tx['merchant']
        current_merchant, current_note = tx['merchant'], None

    m = row_re.search(tx.get('originalText', ''))
    row_num = int(m.group(1)) if m else None
    formula = amount_formulas.get(row_num) if row_num else None

    rows.append({
        'id': tx['id'],
        'text': base_text,
        'formula': formula,
        'current_merchant': current_merchant,
        'current_note': current_note,
        'amount': tx['amount'],
        'type': tx['type'],
    })

FEW_SHOT = """
判斷範例（正確拆法）：
1. "戀與製作人-特權卡" -> merchant="戀與製作人", items=[], note="特權卡"
   （遊戲儲值/商店類：即使銀行帳單顯示的是Google/MyCard等付款通道，商家欄位一律填
   真正的服務/商品本體名稱，這裡是遊戲名，不是付款通道名）
2. "《戀與製作人》週年饋禮" -> merchant="戀與製作人", items=[{"name":"週年饋禮"}], note=null
3. "預購-酷比樂-LOOK UP 名偵探柯南 江戶川柯南＆怪盜基德 套組 附特典" ->
   merchant="酷比樂", items=[{"name":"LOOK UP 名偵探柯南 江戶川柯南＆怪盜基德 套組 附特典"}], note="預購"
   （"預購"是購買狀態詞不是商家，但值得保留在 note 裡）
4. "王牌映畫-吉拿棒、熱狗堡" -> merchant="王牌映畫", items=[{"name":"吉拿棒"},{"name":"熱狗堡"}], note=null
5. "BABY MMMM娃娃-銀虎、河玟-0313_beomgyutxt代購（本金）" ->
   merchant="BABY MMMM娃娃", items=[{"name":"銀虎"},{"name":"河玟"}], note="0313批次-本金",
   specialTag={"type":"proxy_purchase","counterparty":"beomgyutxt","note":"0313批次"}
6. "廖妤甄-IKEA餐廳（LINE PAY MONEY湊$120）" -> merchant="IKEA餐廳", items=[],
   note="廖妤甄-LINE PAY MONEY湊$120", specialTag=null
   （人名不是商家；但這種朋友湊錢/分帳的情況目前先不用 specialTag 標記，只需要把人名
   移出 merchant，因為完整的分帳追蹤是之後才要做的獨立功能）
7. "一沐日"（單純商家名，沒有其他資訊，也沒有formula）-> merchant="一沐日", items=[], note=null,
   specialTag=null （這種不用改，維持原樣）

判斷範例（帶公式，公式數字要對應到品項單價）：
8. text="mamamiya-《時光代理人》掛件盲盒、三麗鷗聯動美味時光收藏卡", formula="=-220-130" ->
   merchant="mamamiya", items=[{"name":"《時光代理人》掛件盲盒","unitPrice":220},
   {"name":"三麗鷗聯動美味時光收藏卡","unitPrice":130}], note=null, confidence=0.9
   （公式裡兩個數字220、130，跟商品名稱剛好兩項對得上，照順序指定）
9. text="Lalaport-My Anime Square柯南快閃餅乾-基德、柯南", formula="=-190*2" ->
   merchant="Lalaport-My Anime Square", items=[{"name":"柯南快閃餅乾-基德","unitPrice":190},
   {"name":"柯南快閃餅乾-柯南","unitPrice":190}], note=null, confidence=0.85
   （*2 代表兩個同價190的品項，都是「柯南快閃餅乾」系列的不同角色款）
10. text="邵瀞葶代購-鐳塔守護神Q版-小卡鐳塔各一、顧江小立牌各一", formula="=ROUNDUP(-(4.2*9+17.5*2)*4.45,0)" ->
    merchant="鐳塔守護神Q版", items=[{"name":"小卡鐳塔","unitPrice":18.69,"quantity":9,"note":"日幣4.2×匯率4.45"},
    {"name":"顧江小立牌","unitPrice":77.88,"quantity":2,"note":"日幣17.5×匯率4.45"}], note=null,
    specialTag={"type":"proxy_purchase","counterparty":"邵瀞葶","note":null}, confidence=0.75
    （公式是 (單價4.2×數量9 + 單價17.5×數量2) 再乘匯率4.45；unitPrice記錄成「換算成台幣後」的
    單價=外幣單價×匯率，note寫換算依據；"各一"其實是"各9個/各2個"更貼近公式，以公式數字為準）
11. text="清心福全", formula="=-60-55" -> merchant="清心福全", items=[], note=null,
    specialTag=null, confidence=0.3
    （公式有兩個數字但文字完全沒列品項名稱，不知道60跟55各是什麼，不要亂猜品項名稱，
    items留空、confidence壓低即可，讓Ivy自己看公式決定要不要手動補）
"""

SYSTEM_PROMPT = f"""你輸出的所有文字（merchant/items/note/specialTag 裡的中文）一律使用繁體中文，
絕對不要出現簡體字，即使原始文字或你自己的知識庫傾向用簡體也要轉換成繁體輸出。

你是財務記帳資料清理專家。使用者「Ivy」的記帳習慣是把商家、買了什麼、
折扣/備註全部寫在同一個欄位裡，有時候她會用 Excel 公式（例如 =-220-130）記錄每樣
商品的單價，現在要把這些拆成乾淨欄位：

- merchant：真正的商家/服務本體名稱（不是購買狀態詞如「預購」「退款」，不是人名，
  不是付款通道如 Google/MyCard——遊戲儲值一律用遊戲本身當商家）
- items：這筆交易買的商品清單（陣列，可能是空陣列）。每個品項是
  {{"name": "...", "unitPrice": 數字或省略, "quantity": 數字或省略, "note": "..."或省略}}。
  只有在附上的公式（如果有）裡的數字能明確對應到這個品項時，才填 unitPrice/quantity；
  對不上、或沒有公式可以佐證，就不要填單價，寧可留空也不要用文字裡的形容詞亂猜金額。
  單價一律填「換算成台幣後」的金額（如果公式裡有乘匯率，unitPrice = 外幣單價×匯率，
  換算依據寫進該品項的 note，例如「日幣4.2×匯率4.45」）。
- note：不屬於任何品項、也不是商家名的其餘說明文字（例如「預購」「已打統編」
  批次代號、朋友分帳的脈絡說明）；沒有就填 null

另外判斷是否符合以下兩種特殊性質（只有明確符合才標記，一般消費不要標）：
- proxy_purchase（代購）：有人幫使用者代買東西，通常會出現「代購」「代轉」字樣，
  且能看出代購人是誰（帳號/暱稱/人名）
- work_advance（工作代墊）：使用者先自己出錢買工作用品，之後要跟公司/主管報帳，
  通常會有「工作」「統編」「報帳」等字樣

符合的話輸出 specialTag: {{"type": "proxy_purchase"|"work_advance", "counterparty": "代購人或報帳對象", "note": "額外說明或null"}}，不符合就輸出 specialTag: null。

{FEW_SHOT}

如果一筆資料看起來已經很乾淨、不需要改，merchant/items/note 就照抄現況，不要為了
改而改。你看不懂、無法判斷商家是誰的，merchant 維持原文字，items 給空陣列，
note 給 null，並把 confidence 設低（0~1，愈不確定愈低）。

輸出格式：純 JSON 陣列，每個元素對應輸入的一筆，順序跟輸入一致，且必須帶著輸入的 id：
[{{"id": "...", "merchant": "...", "items": [{{"name":"...", "unitPrice":數字或省略, "quantity":數字或省略, "note":"..."或省略}}, ...],
  "note": "..." or null,
  "specialTag": null or {{"type":"...", "counterparty":"...", "note": "..." or null}},
  "confidence": 0.0~1.0}}, ...]
不要輸出任何 JSON 以外的文字。
"""

MODEL = 'gemini-3-flash-preview'
# 免費額度是「每天每個模型20次請求」，750筆如果一批25筆要送30次，一天額度不夠用。
# 改成大批次（100筆/次）壓在 8 次以內，留緩衝給重試。
BATCH_SIZE = 100
MAX_RETRIES = 2
PROGRESS_PATH = os.path.join(HERE, 'reparse_progress.json')


def call_gemini(client, batch):
    payload = []
    for r in batch:
        item = {'id': r['id'], 'text': r['text'], 'amount': r['amount'], 'type': r['type']}
        if r['formula']:
            item['formula'] = r['formula']
        payload.append(item)
    prompt = SYSTEM_PROMPT + '\n\n待判斷資料：\n' + json.dumps(payload, ensure_ascii=False)
    for attempt in range(MAX_RETRIES):
        try:
            resp = client.models.generate_content(
                model=MODEL,
                contents=prompt,
                config={'response_mime_type': 'application/json'},
            )
            text = resp.text
            if not text:
                raise RuntimeError('empty response')
            text = re.sub(r'^```json\s*|\s*```$', '', text.strip())
            return json.loads(text)
        except Exception as e:
            print(f'  batch failed (attempt {attempt + 1}): {e}')
            if attempt < MAX_RETRIES - 1:
                time.sleep(10)
    return None  # 這批失敗，不中斷整支腳本，留給下次重跑（有checkpoint，不會重算已完成的批次）


def load_progress():
    if os.path.exists(PROGRESS_PATH):
        with open(PROGRESS_PATH, encoding='utf-8') as f:
            return json.load(f)
    return []


def save_progress(results):
    with open(PROGRESS_PATH, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)


def main():
    client = genai.Client(api_key=load_api_key())
    rows_by_id = {r['id']: r for r in rows}

    all_results = load_progress()
    done_ids = {r['id'] for r in all_results}
    remaining = [r for r in rows if r['id'] not in done_ids]
    print(f'{len(done_ids)} rows already done (checkpoint), {len(remaining)} remaining.')

    stopped_early = False
    for i in range(0, len(remaining), BATCH_SIZE):
        batch = remaining[i:i + BATCH_SIZE]
        print(f'processing batch {i + 1}-{i + len(batch)} / {len(remaining)} remaining ...')
        results = call_gemini(client, batch)
        if results is None:
            print('batch failed, stopping here (checkpoint saved, safe to re-run later to resume).')
            stopped_early = True
            break
        all_results.extend(results)
        save_progress(all_results)

    corrections = []
    review_lines = []
    low_confidence_lines = []

    for res in all_results:
        r = rows_by_id.get(res.get('id'))
        if not r:
            continue
        new_merchant = (res.get('merchant') or '').strip()
        raw_items = res.get('items') or []
        new_items = []
        for it in raw_items:
            name = (it.get('name') or '').strip()
            if not name:
                continue
            entry = {'name': name}
            if it.get('unitPrice') is not None:
                entry['unitPrice'] = it['unitPrice']
            if it.get('quantity') is not None:
                entry['quantity'] = it['quantity']
            if it.get('note'):
                entry['note'] = it['note']
            new_items.append(entry)
        new_note = (res.get('note') or '').strip() or None
        special_tag = res.get('specialTag')
        confidence = res.get('confidence', 1)

        changed = (
            new_merchant != r['current_merchant']
            or bool(new_items)
            or new_note != r['current_note']
            or special_tag is not None
        )
        if not changed:
            continue

        entry = {
            'id': r['id'],
            'old_merchant': r['current_merchant'],
            'old_note': r['current_note'],
            'merchant': new_merchant or r['current_merchant'],
            'items': new_items or None,
            'note': new_note,
            'specialTag': special_tag,
        }
        corrections.append(entry)

        line = (f"{r['id'][:8]} | 原:{r['current_merchant']!r} note={r['current_note']!r}"
                f"{' formula=' + r['formula'] if r['formula'] else ''}\n"
                f"  -> merchant={entry['merchant']!r} items={entry['items']} note={entry['note']!r} "
                f"specialTag={entry['specialTag']} (confidence={confidence})")
        review_lines.append(line)
        if confidence is not None and confidence < 0.6:
            low_confidence_lines.append(line)

    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(corrections, f, ensure_ascii=False, indent=2)

    with open(REVIEW_PATH, 'w', encoding='utf-8') as f:
        if stopped_early:
            f.write(f'!! 尚未跑完全部775筆（Gemini額度限制中斷），已處理 {len(all_results)}/{len(rows)} 筆。\n')
            f.write('!! 已完成的部分有checkpoint（reparse_progress.json），重新執行這支腳本會自動接著跑剩下的，不會重算。\n\n')
        f.write(f'重新解析、建議修改的筆數: {len(corrections)} / {len(all_results)} 已處理\n')
        f.write(f'原始xlsx裡金額欄位是公式的筆數: {len(amount_formulas)}\n\n')
        for line in review_lines:
            f.write(line + '\n')
        f.write(f'\n\n信心度偏低(<0.6)，建議 Ivy 優先看這些 ({len(low_confidence_lines)}):\n')
        for line in low_confidence_lines:
            f.write(line + '\n')

    print(f'done. processed {len(all_results)}/{len(rows)}, corrections: {len(corrections)}, low confidence: {len(low_confidence_lines)}, stopped_early={stopped_early}')


if __name__ == '__main__':
    main()

# -*- coding: utf-8 -*-
# 2026-07-23 全面資料查核 Phase 1：現金支出2023/2024共889筆(含1筆分裝多出來的子項目，
# 原始row數888)，逐row重新讀取原始xlsx(「2023」「2024」兩個分頁)比對日期/金額，
# 不依賴任何中間產物。originalText保留了原始sheet名稱+row編號
# ("現金支出日記帳匯入 sheet2024 row123")，可以逐row精確比對。
#
# 欄位：A=日期(當月第幾天，float，blank=延續上一列同一天；月份標題列A欄是中文月份字串)、
# B=支付(付款方式，跟金額/日期查核無關)、C=分類、D=地點、E=項目、F=價錢(可能是formula-like
# 文字，如"59+35")、G=備註。月份標題列(C/D/E全空、F是當月月結總額)不是交易，要跳過但仍佔一個
# row編號。
import json, re, sys, io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = r'C:\Users\Master\Projects\Pawket'
XLSX_PATH = ROOT + r'\對帳資料\現金支出與悠遊卡\2023-2024現金支出.xlsx'
UNIFIED_PATH = ROOT + r'\Pawket\匯入_統整全部.json'

MONTH_MAP = {'一月': 1, '二月': 2, '三月': 3, '四月': 4, '五月': 5, '六月': 6,
             '七月': 7, '八月': 8, '九月': 9, '十月': 10, '十一月': 11, '十二月': 12}


def parse_price(v):
    # 2026-07-23查證：價錢欄位有些是純數字文字，有些前面帶#(例如'#150')，
    # 意義跟純數字一樣，只是Ivy記帳習慣的寫法，不是格式錯誤。
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lstrip('#')
    if re.match(r'^-?[0-9]+(\.[0-9]+)?$', s):
        return float(s)
    cleaned = s.replace(' ', '')
    if re.match(r'^[0-9+\-*/(). ]+$', cleaned) and re.search(r'[0-9]', cleaned):
        try:
            return float(eval(cleaned, {'__builtins__': {}}, {}))
        except Exception:
            return None
    return None


wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

source_by_sheet_row = {}
month_header_totals = []
for sheet_name, year in [('2023', 2023), ('2024', 2024)]:
    ws = wb[sheet_name]
    cur_month = None
    cur_year = year
    cur_day = None
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        cat = ws.cell(row=r, column=3).value
        place = ws.cell(row=r, column=4).value
        item = ws.cell(row=r, column=5).value
        price_v = ws.cell(row=r, column=6).value
        if isinstance(a, str) and a in MONTH_MAP and cat is None and place is None and item is None:
            new_month = MONTH_MAP[a]
            # 年度跨年：這個分頁的資料如果月份數字比上一次記錄的月份還小(例如12月後又出現1月)，
            # 代表跨到下一年了(例如「2024」分頁尾端其實延伸到2025年1月初)，年份要跟著+1。
            if cur_month is not None and new_month < cur_month:
                cur_year += 1
            cur_month = new_month
            cur_day = None
            month_header_totals.append((sheet_name, r, a, price_v))
            continue
        if a is None and cat is None and place is None and item is None and price_v is None:
            continue
        # 年度總計列(2024第2列)：A/cat/place/item全空但price有值，且還沒遇到任何月份標題
        if cur_month is None:
            continue
        if isinstance(a, (int, float)):
            cur_day = int(a)
        if cur_month is None or cur_day is None:
            continue
        date_str = f'{cur_year}-{cur_month:02d}-{cur_day:02d}'
        price = parse_price(price_v)
        has_content = cat is not None or place is not None or item is not None
        source_by_sheet_row[(sheet_name, r)] = {'date': date_str, 'amount': price, 'has_content': has_content}

with open(UNIFIED_PATH, encoding='utf-8') as f:
    data = json.load(f)
txs = data['transactions'] if isinstance(data, dict) else data
cash = [t for t in txs if (t.get('originalText') or '').startswith('現金支出日記帳匯入')]

real_mismatches = []
blank_source_price = []  # 原始xlsx這格價錢是空白，系統卻有金額——可能是Ivy事後手動補的，不是bug，但列出來供人工核對
mismatches = []
matched = set()
for t in cash:
    m = re.match(r'現金支出日記帳匯入 sheet(\d+) row(\d+)', t['originalText'])
    if not m:
        mismatches.append(('NO_ROW_TAG', t.get('id'), t.get('originalText')))
        continue
    sheet_name, row = m.group(1), int(m.group(2))
    key = (sheet_name, row)
    matched.add(key)
    src = source_by_sheet_row.get(key)
    if src is None:
        mismatches.append(('ROW_NOT_FOUND_IN_SOURCE', t.get('id'), key, t.get('merchant'), t.get('amount')))
        continue
    src_amount = abs(src['amount']) if src['amount'] is not None else None
    if src_amount is None:
        blank_source_price.append((key, t.get('merchant'), 'system=', t['amount']))
    elif round(src_amount, 2) != round(t['amount'], 2):
        real_mismatches.append(('AMOUNT_MISMATCH', key, t.get('merchant'), 'system=', t['amount'], 'source=', src_amount, t.get('originalText')))
    if src['date'] != t['date']:
        real_mismatches.append(('DATE_MISMATCH', key, t.get('merchant'), 'system=', t['date'], 'source=', src['date']))

missing_in_system = [k for k in source_by_sheet_row if k not in matched and source_by_sheet_row[k]['has_content']]

print(f'系統裡現金支出來源紀錄共 {len(cash)} 筆')
print(f'原始xlsx解析出的有效資料row共 {len(source_by_sheet_row)} 筆')
print(f'月份標題列共 {len(month_header_totals)} 個 (僅供資訊，不參與逐筆比對):')
for mh in month_header_totals:
    print(' ', mh)
print(f'其他問題(NO_ROW_TAG等) {len(mismatches)} 筆:')
for m in mismatches:
    print(' ', m)
print(f'\\n=== 真正的金額/日期不一致(需要人工檢查) 共 {len(real_mismatches)} 筆 ===')
for m in real_mismatches:
    print(' ', m)
print(f'\\n=== 原始xlsx價錢欄空白、但系統有金額(可能是Ivy事後手動補的，非bug) 共 {len(blank_source_price)} 筆 ===')
for m in blank_source_price:
    print(' ', m)
print(f'\\n=== 原始xlsx裡有內容、但系統完全沒有對應紀錄的row 共 {len(missing_in_system)} 筆 ===')
for k in sorted(missing_in_system):
    print(' ', k, source_by_sheet_row[k])

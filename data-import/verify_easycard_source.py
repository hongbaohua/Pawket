# -*- coding: utf-8 -*-
# 2026-07-23 全面資料查核 Phase 1：二技悠遊卡39筆，逐row比對原始xlsx(「二技悠遊卡餘額」分頁)
# 的日期/金額，同時用「餘額」欄位做running balance交叉驗證(上一列餘額+這列金額=這列餘額)，
# 抓比單純比對日期/金額更嚴格的錯誤(例如金額正負號錯誤但剛好...其實不太可能同時通過兩種
# 檢查，兩者一起做更保險)。
import json, re, sys, io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


def parse_amount(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if re.match(r'^-?[0-9]+(\.[0-9]+)?$', s):
        return float(s)
    cleaned = s.replace(' ', '')
    if re.match(r'^[0-9+\-*/(). ]+$', cleaned):
        return float(eval(cleaned, {'__builtins__': {}}, {}))
    return None

ROOT = r'C:\Users\Master\Projects\Pawket'
XLSX_PATH = ROOT + r'\對帳資料\現金支出與悠遊卡\2023-2024現金支出.xlsx'
UNIFIED_PATH = ROOT + r'\Pawket\匯入_統整全部.json'

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb['二技悠遊卡餘額']

source_by_row = {}
last_date = None
prev_balance = None
balance_mismatches = []
for r in range(2, ws.max_row + 1):
    date_v = ws.cell(row=r, column=1).value
    place = ws.cell(row=r, column=2).value
    purpose = ws.cell(row=r, column=3).value
    amount_v = ws.cell(row=r, column=4).value
    balance_v = ws.cell(row=r, column=5).value
    if date_v is None and place is None and purpose is None and amount_v is None:
        continue
    if date_v is not None:
        if hasattr(date_v, 'isoformat'):
            last_date = date_v.isoformat()[:10]
        else:
            s = str(int(date_v))
            last_date = f'{s[0:4]}-{s[4:6]}-{s[6:8]}'
    amount_num = parse_amount(amount_v)
    balance_num = parse_amount(balance_v)
    if balance_num is not None and amount_num is not None and prev_balance is not None:
        expected = round(prev_balance + amount_num, 2)
        if round(balance_num, 2) != expected:
            balance_mismatches.append((r, 'expected_balance=', expected, 'actual=', balance_num))
    if balance_num is not None:
        prev_balance = balance_num
    source_by_row[r] = {'date': last_date, 'amount': amount_num}

with open(UNIFIED_PATH, encoding='utf-8') as f:
    data = json.load(f)
txs = data['transactions'] if isinstance(data, dict) else data
ec = [t for t in txs if (t.get('originalText') or '').startswith('二技悠遊卡餘額分頁匯入')]

mismatches = []
matched_rows = set()
for t in ec:
    m = re.match(r'二技悠遊卡餘額分頁匯入 row(\d+)', t['originalText'])
    if not m:
        mismatches.append(('NO_ROW_TAG', t.get('id'), t.get('originalText')))
        continue
    row = int(m.group(1))
    matched_rows.add(row)
    src = source_by_row.get(row)
    if src is None:
        mismatches.append(('ROW_NOT_FOUND_IN_SOURCE', t.get('id'), row, t.get('merchant'), t.get('amount')))
        continue
    src_amount = abs(src['amount']) if src['amount'] is not None else None
    if src_amount is None or round(src_amount, 2) != round(t['amount'], 2):
        mismatches.append(('AMOUNT_MISMATCH', row, t.get('merchant'), 'system=', t['amount'], 'source=', src_amount))
    if src['date'] != t['date']:
        mismatches.append(('DATE_MISMATCH', row, t.get('merchant'), 'system=', t['date'], 'source=', src['date']))

missing_in_system = [r for r in source_by_row if r not in matched_rows]
# 原始表格最後2列(2025-12-27附近)5.8節已確認判斷不匯入是對的，這裡預期會出現在missing清單，不算bug
print(f'系統裡二技悠遊卡來源紀錄共 {len(ec)} 筆')
print(f'原始xlsx解析出的有效資料row共 {len(source_by_row)} 筆')
print(f'running balance自我檢查(上一列餘額+金額=這一列餘額) mismatch共 {len(balance_mismatches)} 筆:')
for b in balance_mismatches:
    print(' ', b)
print(f'系統vs來源 mismatch共 {len(mismatches)} 筆:')
for m in mismatches:
    print(' ', m)
print(f'原始xlsx裡有、但系統完全沒有對應紀錄的row共 {len(missing_in_system)} 筆:')
for r in sorted(missing_in_system):
    print(' ', r, source_by_row[r])

# -*- coding: utf-8 -*-
# 2026-07-23 全面資料查核 Phase 1：重新讀取最原始的中信xlsx，逐row跟目前的
# 匯入_統整全部.json比對日期/金額，不依賴任何中間產物，確認fix_001~008這一整層
# 修正沒有意外改壞任何一筆的原始金額/日期（見PROJECT_STATUS.md第5.14節）。
#
# 沿用build_import2.py已經驗證過的解析邏輯(safe_eval/日期carry-forward)，
# 因為originalText保留了原始xlsx的row編號("中信對帳單匯入 row123 | type=XXX")，
# 可以逐row精確比對，不用重建整個月結checksum。
import json, re, math, sys, io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = r'C:\Users\Master\Projects\Pawket'
XLSX_PATH = ROOT + r'\對帳資料\中國信託\中信餘額_正確答案.xlsx'
UNIFIED_PATH = ROOT + r'\Pawket\匯入_統整全部.json'


def safe_eval(formula):
    expr = formula.lstrip('=')
    negate_outer = False
    if expr.upper().startswith('-ROUND'):
        negate_outer = True
        expr = expr[1:]
    m = re.match(r'^ROUND(UP)?\((.+),\s*(\d+)\)$', expr, re.IGNORECASE)
    if m:
        inner, digits = m.group(2), int(m.group(3))
        inner = inner.replace('%', '/100')
        val = eval(inner, {'__builtins__': {}}, {})
        factor = 10 ** digits
        if m.group(1):
            result = math.ceil(val * factor) / factor if val >= 0 else -math.ceil(-val * factor) / factor
        else:
            result = round(val, digits)
        return -result if negate_outer else result
    cleaned = expr.replace('%', '/100')
    if not re.match(r'^[0-9+\-*/(). ]+$', cleaned):
        return None
    return eval(cleaned, {'__builtins__': {}}, {})


wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
ws = wb.active
wbv = openpyxl.load_workbook(XLSX_PATH, data_only=True)
wsv = wbv.active

source_by_row = {}
last_date = None
for r in range(2, ws.max_row + 1):
    date_v = wsv.cell(row=r, column=1).value
    ttype = wsv.cell(row=r, column=2).value
    amount = wsv.cell(row=r, column=4).value
    raw_d = ws.cell(row=r, column=4).value
    if amount is None and isinstance(raw_d, str) and raw_d.startswith('='):
        amount = safe_eval(raw_d)
    merchant = ws.cell(row=r, column=3).value
    if merchant is None and amount is None and ttype is None:
        continue
    if ttype == '未登錄':
        continue
    if date_v is not None:
        last_date = date_v
    date_str = last_date.isoformat()[:10] if last_date is not None else None
    source_by_row[r] = {'date': date_str, 'type': ttype, 'amount': amount}

with open(UNIFIED_PATH, encoding='utf-8') as f:
    data = json.load(f)
txs = data['transactions'] if isinstance(data, dict) else data
ctbc = [t for t in txs if (t.get('originalText') or '').startswith('中信對帳單匯入')]

mismatches = []
matched_rows = set()
for t in ctbc:
    m = re.match(r'中信對帳單匯入 row(\d+)', t['originalText'])
    if not m:
        mismatches.append(('NO_ROW_TAG', t.get('id'), t.get('originalText')))
        continue
    row = int(m.group(1))
    matched_rows.add(row)
    src = source_by_row.get(row)
    if src is None:
        mismatches.append(('ROW_NOT_FOUND_IN_SOURCE', t.get('id'), row, t.get('merchant'), t.get('amount')))
        continue
    src_amount = abs(src['amount']) if src['amount'] is not None else None
    if src_amount is None or round(src_amount, 2) != round(t['amount'], 2):
        mismatches.append(('AMOUNT_MISMATCH', row, t.get('merchant'), 'system=', t['amount'], 'source=', src_amount))
    if src['date'] != t['date']:
        mismatches.append(('DATE_MISMATCH', row, t.get('merchant'), 'system=', t['date'], 'source=', src['date']))

# 反向檢查：source有的row，系統裡完全沒有對應紀錄(代表匯入時漏掉了)
missing_in_system = [r for r in source_by_row if r not in matched_rows]

print(f'系統裡中信來源紀錄共 {len(ctbc)} 筆')
print(f'原始xlsx解析出的有效資料row共 {len(source_by_row)} 筆')
print(f'比對到的mismatch共 {len(mismatches)} 筆:')
for m in mismatches:
    print(' ', m)
print(f'原始xlsx裡有、但系統完全沒有對應紀錄的row共 {len(missing_in_system)} 筆:')
for r in sorted(missing_in_system):
    print(' ', r, source_by_row[r])

# -*- coding: utf-8 -*-
# 2026-07-23 全面資料查核 Phase 2：把全部交易依照系統的完整欄位匯出成Excel，給Ivy逐筆審核。
# 2026-07-24 改版：
# 1. 資料來源改成直接查詢Supabase live資料(用query_supabase.py，SUPABASE_SERVICE_ROLE_KEY
#    已經存在.env.local)，不再讀本機的`匯入_統整全部.json`——那份檔案是匯入當下的快照，
#    後續fix_001~010.sql這些修正如果本機json忘記同步就會跟資料庫真正的內容不一致，
#    直接查live資料庫可以徹底避免這個問題，也能反映Ivy自己在App裡動手改過的任何資料。
# 2. 新增accountId欄位(原本漏掉，只有轉帳用的fromAccountId/toAccountId，一般收支
#    完全沒有顯示是哪個帳戶)，帳戶一律顯示名稱(不是id)，方便Ivy閱讀/填寫。
# 3. 新增第二個分頁「新增交易範本」：Ivy要新增的交易可以直接照這個範本填在這個分頁，
#    第2列(淺綠底)是每個欄位的填寫說明，第3列起是空白列給她填。
#    她的計畫：這次查核完(改「全面資料查核」分頁+填「新增交易範本」分頁)之後，
#    要把資料庫全部清掉，直接照這份Excel(修正過的既有資料+新增的資料)重新匯入，
#    所以這份檔案要能同時當作「查核紀錄」跟「下一次匯入的完整來源」兩種用途。
import sys, os, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from query_supabase import fetch, fetch_all

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = r'C:\Users\Master\Projects\Pawket'
OUT_PATH = ROOT + r'\Pawket\data-import\全面資料查核表_2026-07-23.xlsx'
# Ivy反應在Pawket\data-import\深層路徑裡找不到這份檔案——比照CLAUDE.md「不該讓Ivy自己
# 巡查資料夾找東西」的原則，另外存一份在最外層根目錄，這才是Ivy真正應該打開的位置，
# 每次重新產生都會同步更新這個副本。
IVY_COPY_PATH = ROOT + r'\查核表_請在這裡填寫.xlsx'
IVY_USER_ID = '56dd1f4e-32c5-41ba-8da4-7eabce8b7b70'  # 見PROJECT_STATUS.md「新能力」小節

print('查詢Supabase即時資料中...')
accounts_raw = fetch('accounts', f'select=id,name&user_id=eq.{IVY_USER_ID}')
ACCOUNT_NAME_BY_ID = {a['id']: a['name'] for a in accounts_raw}
print(f'帳戶共{len(ACCOUNT_NAME_BY_ID)}個')

rows_raw = fetch_all(
    'transactions',
    f'select=id,date,merchant,type,net_amount,gross_amount,discounts,l1,l2,l3,'
    f'account_id,payment_channel,items,note,special_tag,from_account_id,to_account_id,'
    f'original_text&user_id=eq.{IVY_USER_ID}&deleted_at=is.null'
)
print(f'交易共{len(rows_raw)}筆')


def acct_name(account_id):
    if not account_id:
        return ''
    return ACCOUNT_NAME_BY_ID.get(account_id, f'(未知帳戶id:{account_id})')


# 轉成跟舊版腳本一樣方便存取的dict結構(欄位名沿用駝峰式，跟types.ts一致)
txs = []
for r in rows_raw:
    txs.append({
        'id': r['id'],
        'date': r['date'],
        'merchant': r['merchant'],
        'type': r['type'],
        'amount': r['net_amount'],
        'grossAmount': r.get('gross_amount'),
        'discounts': r.get('discounts'),
        'category': {'l1': r.get('l1') or '', 'l2': r.get('l2') or '', 'l3': r.get('l3') or ''},
        'accountId': r.get('account_id'),
        'paymentChannel': r.get('payment_channel'),
        'items': r.get('items'),
        'note': r.get('note'),
        'specialTag': r.get('special_tag'),
        'fromAccountId': r.get('from_account_id'),
        'toAccountId': r.get('to_account_id'),
        'originalText': r.get('original_text'),
    })

# ============ 這次Phase 2查證後，確定要改的欄位修正(2026-07-23那輪查核結果) ============
CONFIRMED_FIXES = {}


def find_one(merchant, date, amount):
    matches = [t for t in txs if t['merchant'] == merchant and t['date'] == date and abs(t['amount'] - amount) < 0.01]
    if len(matches) != 1:
        print(f'WARNING: {merchant} {date} {amount} 找到{len(matches)}筆，預期1筆', file=sys.stderr)
        return None
    return matches[0]['id']


fixes_applied = {}

t = find_one('波妮國際', '2023-10-23', 1161)
if t:
    fixes_applied[t] = {
        'l1': 'Variable', 'l2': '服飾美妝', 'l3': '',
        '備註': 'WebSearch查證「波妮國際有限公司」是登記在台中北區的內衣零售業公司，原分類「其他雜項」改成「服飾美妝」。'
    }

t = find_one('樂士Luxe3C', '2026-05-21', 390)
if t:
    fixes_applied[t] = {
        'l1': 'Variable', 'l2': '3C電子', 'l3': '',
        '備註': '品項是「Type-C轉接線」，明顯是3C電子配件，原分類「服飾美妝」看起來是誤植，改成「3C電子」。'
    }

t = find_one('中友', '2025-09-25', 580)
if t:
    fixes_applied[t] = {
        'l1': 'Variable', 'l2': '休閒娛樂', 'l3': '玩具',
        '備註': '品項是「柯南盲盒×2」，跟同一個商家「中友」其他7筆柯南盲盒紀錄都歸「休閒娛樂/玩具」不一致，這筆原本是「生活日用/百貨公司」，改成跟其他7筆一致。'
    }

t = find_one('先喝道', '2026-03-26', 65)
if t:
    fixes_applied[t] = {
        'l1': 'Variable', 'l2': '餐飲食品', 'l3': '飲料',
        '備註': 'WebSearch查證「先喝道(TAOTAOTEA)」是古典玫瑰園集團旗下的手搖飲品牌，不是電影票，原分類「休閒娛樂/電影票」改成「餐飲食品/飲料」。'
    }

for merchant, date, amount in [('速風達', '2026-07-03', 774), ('速風達', '2026-07-05', 142)]:
    t = find_one(merchant, date, amount)
    if t:
        fixes_applied[t] = {
            'specialTag': {'type': 'proxy_purchase', 'counterparty': '(待Ivy補充代購對象)', 'note': None},
            '備註': '備註欄寫「代購」/「代購運費」但沒有設定代購性質標記(specialTag)，補上——代購對象欄位需要Ivy補充是誰代購的。'
        }

# ============ 軟性建議/需要Ivy補充資訊的項目，只加備註不改欄位 ============
SOFT_NOTES = {}

t = find_one('知翎文化', '2024-11-30', 957)
if t:
    SOFT_NOTES[t] = '品項是《時光代理人》美術設定集+運費，性質比較像收藏品，軟性建議「其他雜項」改成「休閒娛樂」，但不確定，請Ivy自己判斷。'

t = find_one('高鐵', '2020-07-24', 500)
if t:
    SOFT_NOTES[t] = 'originalText顯示「高鐵智慧型手機Android」，商家/描述混雜不清楚，懷疑是原始PDF對帳單解析時把兩筆資訊黏在一起，Claude Code查不出這筆實際上是什麼消費，請Ivy自己回憶或查證。'

t = find_one('匯款', '2024-03-16', 500)
if t:
    SOFT_NOTES[t] = '品項是PLAVE迷你二輯空專×5，分類「休閒娛樂/專輯」是對的，但商家欄位填的是付款方式「匯款」不是真正商家，建議改成代購/團購主揪的名字或平台名稱，請Ivy補充實際是跟誰/哪個平台買的。'

t = find_one('郵局', '2024-07-15', 216)
if t:
    SOFT_NOTES[t] = '品項是《戀與製作人》珍藏卡×2件，分類「休閒娛樂/收藏卡」是對的，但商家欄位填的是取貨地點「郵局」不是真正商家，建議改成實際購買的商家/代購對象，請Ivy補充。'

t = find_one('統一超商', '2024-11-06', 182)
if t:
    SOFT_NOTES[t] = '同樣是超商消費，這筆分類「生活日用」，但其他7-11/全家的紀錄多半分類「餐飲食品」，可能只是這筆買的東西剛好不是吃的，不確定，列出來供參考，不算錯誤。'

# ============ 套用確定的修正到記憶體中的資料(還沒寫回資料庫，等Ivy看過表格確認) ============
for tx in txs:
    if tx['id'] in fixes_applied:
        fx = fixes_applied[tx['id']]
        if 'l1' in fx:
            tx['category'] = {'l1': fx['l1'], 'l2': fx['l2'], 'l3': fx['l3']}
        if 'specialTag' in fx:
            tx['specialTag'] = fx['specialTag']

print(f'套用了 {len(fixes_applied)} 筆確定修正')
print(f'加了 {len(SOFT_NOTES)} 筆軟性建議備註')


# ============ 共用格式化函式 ============
def fmt_items(items):
    if not items:
        return ''
    parts = []
    for it in items:
        s = it.get('name', '')
        if it.get('unitPrice') is not None:
            s += f"(${it['unitPrice']}"
            if it.get('quantity') and it['quantity'] != 1:
                s += f"×{it['quantity']}"
            s += ')'
        elif it.get('quantity') and it['quantity'] != 1:
            s += f"×{it['quantity']}"
        if it.get('note'):
            s += f"[{it['note']}]"
        parts.append(s)
    return '; '.join(parts)


def fmt_discounts(discounts):
    if not discounts:
        return ''
    return '; '.join(f"{d['label']}:-${d['amount']}" for d in discounts)


COLUMNS = ['id', 'date', 'merchant', 'type', 'amount', 'grossAmount', 'discounts',
           'l1', 'l2', 'l3', 'accountId', 'paymentChannel', 'items', 'note',
           'specialTag_type', 'specialTag_counterparty', 'specialTag_note',
           'fromAccountId', 'toAccountId', 'originalText', 'AI查核備註', 'Ivy的備註']
IVY_NOTE_COL = len(COLUMNS)
WIDTHS = [10, 11, 18, 9, 8, 10, 16, 10, 10, 10, 12, 12, 30, 20, 12, 14, 14, 12, 12, 40, 55, 40]

L1_LABEL = {'Fixed': '固定支出', 'Variable': '變動支出', 'Investment': '投資儲蓄', 'Income': '收入帳戶'}

header_font = Font(name='Arial', bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
normal_font = Font(name='Arial', size=10)
remark_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
ivy_note_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

wb = openpyxl.Workbook()

# ============ 分頁1：全面資料查核 ============
ws = wb.active
ws.title = '全面資料查核'

for ci, col in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

row_i = 2
for tx in sorted(txs, key=lambda t: (t['date'], t['id'])):
    st = tx.get('specialTag') or {}
    remark = ''
    if tx['id'] in fixes_applied:
        remark = '[已修正] ' + fixes_applied[tx['id']]['備註']
    elif tx['id'] in SOFT_NOTES:
        remark = '[請Ivy確認] ' + SOFT_NOTES[tx['id']]

    values = [
        tx['id'], tx['date'], tx['merchant'], tx['type'], tx['amount'],
        tx.get('grossAmount', ''), fmt_discounts(tx.get('discounts')),
        L1_LABEL.get(tx['category']['l1'], tx['category']['l1']), tx['category']['l2'], tx['category'].get('l3', ''),
        acct_name(tx.get('accountId')), tx.get('paymentChannel', ''), fmt_items(tx.get('items')), tx.get('note', ''),
        st.get('type', ''), st.get('counterparty', ''), st.get('note', ''),
        acct_name(tx.get('fromAccountId')), acct_name(tx.get('toAccountId')),
        tx.get('originalText', ''), remark, '',
    ]
    for ci, v in enumerate(values, start=1):
        cell = ws.cell(row=row_i, column=ci, value=v)
        cell.font = normal_font
        if ci == len(COLUMNS) - 1 and remark:
            cell.fill = remark_fill
        if ci == IVY_NOTE_COL:
            cell.fill = ivy_note_fill
    row_i += 1

for ci, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{row_i - 1}'

# ============ 分頁2：新增交易範本 ============
# Ivy要新增的交易直接照這個範本填在這個分頁，第2列是每個欄位的填寫說明(淺綠底)，
# 第3列起是空白列。她確認完兩個分頁都沒問題後，會把資料庫全部清掉，直接照這份
# Excel(全面資料查核修正過的既有資料 + 新增交易範本裡新填的資料)重新匯入。
ws2 = wb.create_sheet('新增交易範本')

EXPLANATIONS = {
    'id': '留空即可，系統會自動產生新的識別碼',
    'date': '日期，格式YYYY-MM-DD，例如2026-07-24',
    'merchant': '商家名稱＝買了什麼東西/用了什麼服務的本體，不是付款方式(例如遊戲儲值要寫遊戲名稱，不是Google Play)',
    'type': 'expense(支出) / income(收入) / transfer(帳戶互轉，例如提款、儲值)，三選一',
    'amount': '實付金額(折扣後最終付的錢；transfer的話就是轉帳金額)',
    'grossAmount': '折扣前的原始金額，選填，沒有折扣可以留空',
    'discounts': '折扣明細，格式「標籤:-$金額」，多筆用「; 」分隔，例如：LINE POINT:-$40; 會員折扣:-$10',
    'l1': '固定支出／變動支出／投資儲蓄／收入帳戶，四選一(income要選收入帳戶；transfer可以留空)',
    'l2': '次分類，要跟l1配對，例如變動支出可選：餐飲食品/生活日用/交通通勤/休閒娛樂/服飾美妝/3C電子/醫療保健/學習進修/社交人情/寵物花費/銀行手續費/轉帳/網路購物/其他雜項',
    'l3': '細項標籤，選填，例如：飲料、拿鐵',
    'accountId': f'這筆錢是哪個帳戶，填帳戶名稱就好。你目前的帳戶：{", ".join(ACCOUNT_NAME_BY_ID.values())}',
    'paymentChannel': '付款通道，選填，例如：VISA、LINE Pay、方便付-街口支付',
    'items': '品項清單，格式「商品名(單價×數量)[備註]」，多筆用「; 」分隔，例如：小卡LASER組($95×4)[原幣$6×4個×匯率4.9×折扣0.8（無條件進位）]',
    'note': '備註，選填',
    'specialTag_type': 'proxy_purchase(代購) / work_advance(工作代墊)，選填，沒有留空',
    'specialTag_counterparty': '代購人是誰/要跟誰報帳，選填(不重要可以留空)',
    'specialTag_note': '代購額外說明，選填，例如：已打統編',
    'fromAccountId': '只有type=transfer才要填，寫轉出的帳戶名稱',
    'toAccountId': '只有type=transfer才要填，寫轉入的帳戶名稱',
    'originalText': '留空即可，這是系統記錄原始來源用的欄位',
    'AI查核備註': '這個分頁不用填，留空',
    'Ivy的備註': '可以寫給自己看的補充說明，或標記這筆還不確定',
}

for ci, col in enumerate(COLUMNS, start=1):
    cell = ws2.cell(row=1, column=ci, value=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

explain_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
explain_font = Font(name='Arial', size=9, italic=True, color='548235')
for ci, col in enumerate(COLUMNS, start=1):
    cell = ws2.cell(row=2, column=ci, value=EXPLANATIONS.get(col, ''))
    cell.font = explain_font
    cell.fill = explain_fill
    cell.alignment = Alignment(wrap_text=True, vertical='top')

# 空白列給Ivy填，先預留40列的格式(她要繼續往下填也不受限制，Excel不會因為超過這個範圍就不能用)
BLANK_ROWS = 40
for r in range(3, 3 + BLANK_ROWS):
    for ci in range(1, len(COLUMNS) + 1):
        ws2.cell(row=r, column=ci).font = normal_font

for ci, w in enumerate(WIDTHS, start=1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[2].height = 60
ws2.freeze_panes = 'A3'
ws2.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}{2 + BLANK_ROWS}'

wb.save(OUT_PATH)
wb.save(IVY_COPY_PATH)
print('已輸出:', OUT_PATH)
print('Ivy版本(請打開這份):', IVY_COPY_PATH)
print('全面資料查核:', row_i - 2, '筆')
print('新增交易範本: 已建立', BLANK_ROWS, '個空白列供填寫')

import openpyxl, re, math, json, sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from categorize_rules import apply_rules, special_case, is_low_confidence

XLSX_PATH = r'C:\Users\Master\Projects\Pawket\對帳資料\中國信託\中信餘額_正確答案.xlsx'
JSON_PATH = r'C:\Users\Master\Projects\Pawket\對帳資料\Ivy手標歷史分類參考\4-9.json'
OUT_DIR = os.path.dirname(os.path.abspath(__file__))

def safe_eval(formula):
    expr = formula.lstrip('=')
    negate_outer = False
    if expr.upper().startswith('-ROUND'):
        negate_outer = True
        expr = expr[1:]
    m = re.match(r'^ROUND(UP)?\((.+),\s*(\d+)\)$', expr, re.IGNORECASE)
    if m:
        inner, digits = m.group(2), int(m.group(3))
        inner = inner.replace('%', '/100')
        val = eval(inner, {'__builtins__': {}}, {})
        factor = 10 ** digits
        if m.group(1):
            result = math.ceil(val * factor) / factor if val >= 0 else -math.ceil(-val * factor) / factor
        else:
            result = round(val, digits)
        return -result if negate_outer else result
    cleaned = expr.replace('%', '/100')
    if not re.match(r'^[0-9+\-*/(). ]+$', cleaned):
        return None
    return eval(cleaned, {'__builtins__': {}}, {})

FULLWIDTH_OPEN = chr(0xFF08)
FULLWIDTH_CLOSE = chr(0xFF09)

def norm(s):
    if not s or not isinstance(s, str):
        return ''
    chars_to_strip = ' \t-()_.' + FULLWIDTH_OPEN + FULLWIDTH_CLOSE
    result = ''.join(c for c in s if c not in chars_to_strip)
    return result.lower()

wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
ws = wb.active
wbv = openpyxl.load_workbook(XLSX_PATH, data_only=True)
wsv = wbv.active

raw_rows = []
last_date = None
for r in range(2, ws.max_row + 1):
    date_v = wsv.cell(row=r, column=1).value
    ttype = wsv.cell(row=r, column=2).value
    merchant = ws.cell(row=r, column=3).value
    amount = wsv.cell(row=r, column=4).value
    raw_d = ws.cell(row=r, column=4).value
    if amount is None and isinstance(raw_d, str) and raw_d.startswith('='):
        amount = safe_eval(raw_d)
    if merchant is None and amount is None and ttype is None:
        continue
    if ttype == '未登錄':
        continue
    if date_v is not None:
        last_date = date_v
    if not isinstance(merchant, str):
        merchant = str(merchant) if merchant is not None else ''
    raw_rows.append({'row': r, 'date': last_date, 'type': ttype, 'merchant': merchant, 'amount': amount})

def classify(rec):
    t = rec['type']
    a = rec['amount']
    if t == '提款':
        return 'transfer', 'CTBC->CASH'
    if t == '存款':
        return 'transfer', 'CASH->CTBC'
    if t in ('轉入', '轉出'):
        return ('income' if a > 0 else 'expense'), None
    if t == '利息':
        return 'income', None
    return ('income' if a > 0 else 'expense'), None

# Ivy 確認過的個別修正：Excel 把「7-11」自動轉成日期格式，商家其實是 7-ELEVEN。
ROW_OVERRIDES = {
    87: {'merchant': '7-11', 'category': ('Variable', '餐飲食品', '便利商店')},
    183: {'merchant': '7-11', 'category': ('Variable', '餐飲食品', '便利商店')},
    686: {'merchant': '7-11', 'category': ('Variable', '餐飲食品', '便利商店')},
    742: {'merchant': '7-11', 'category': ('Variable', '餐飲食品', '便利商店')},
    99: {'merchant': 'Relove', 'category': ('Variable', '服飾美妝', '除毛膏')},
    513: {'merchant': '蒔初', 'category': ('Variable', '餐飲食品', '餐廳')},
    715: {'merchant': '手機螢幕維修', 'category': ('Variable', '生活日用', '手機維修'),
          'note': '費用內含代墊轉給姊姊的部分；LINE Pay Money原有餘額$300'},
}

for rec in raw_rows:
    if rec['row'] in ROW_OVERRIDES:
        ov = ROW_OVERRIDES[rec['row']]
        rec['merchant'] = ov['merchant']
        rec['override_category'] = ov['category']
        rec['note'] = ov.get('note')
    nature, transfer_dir = classify(rec)
    rec['nature'] = nature
    rec['transfer_dir'] = transfer_dir

with open(JSON_PATH, 'r', encoding='utf-8') as f:
    old_data = json.load(f)
old_txs = old_data['transactions']

from collections import Counter
merchant_cat_counter = {}
for t in old_txs:
    m = norm(t['merchant'])
    if not m:
        continue
    cat_key = (t['category']['l1'], t['category']['l2'], t['category'].get('l3', ''))
    merchant_cat_counter.setdefault(m, Counter())[cat_key] += 1

merchant_to_cat = {m: c.most_common(1)[0][0] for m, c in merchant_cat_counter.items()}
old_merchants_norm = list(merchant_to_cat.keys())

def find_match(merchant_text):
    nm = norm(merchant_text)
    if not nm:
        return None
    if nm in merchant_to_cat:
        return merchant_to_cat[nm], 'exact'
    candidates = [om for om in old_merchants_norm if len(om) >= 2 and (om in nm or nm in om)]
    if candidates:
        best = max(candidates, key=len)
        return merchant_to_cat[best], 'fuzzy:' + best
    return None

still_unresolved = []
final_records = []
for rec in raw_rows:
    if 'override_category' in rec:
        cat = rec['override_category']
        rec['category'] = {'l1': cat[0], 'l2': cat[1], 'l3': cat[2]}
        rec['source'] = 'ivy-confirmed'
        final_records.append(rec)
        continue
    sc = special_case(rec['merchant'], rec['amount'])
    if sc == 'TRANSFER_CTBC_TO_CASH':
        rec['nature'] = 'transfer'
        rec['transfer_dir'] = 'CTBC->CASH'
        rec['category'] = {'l1': 'Variable', 'l2': '轉帳', 'l3': ''}
        rec['source'] = 'special-case-transfer'
        final_records.append(rec)
        continue
    if rec['nature'] == 'transfer':
        rec['category'] = {'l1': 'Variable', 'l2': '轉帳', 'l3': ''}
        rec['source'] = 'transfer-rule'
        final_records.append(rec)
        continue
    if sc is not None:
        rec['category'] = {'l1': sc[0], 'l2': sc[1], 'l3': sc[2]}
        rec['source'] = 'special-case'
        final_records.append(rec)
        continue
    result = find_match(rec['merchant'])
    if result:
        cat, how = result
        rec['category'] = {'l1': cat[0], 'l2': cat[1], 'l3': cat[2]}
        rec['source'] = '4-9.json:' + how
        final_records.append(rec)
        continue
    rule_cat = apply_rules(rec['merchant'])
    if rule_cat:
        rec['category'] = {'l1': rule_cat[0], 'l2': rule_cat[1], 'l3': rule_cat[2]}
        rec['source'] = 'keyword-rule-lowconf' if is_low_confidence(rec['merchant']) else 'keyword-rule'
        final_records.append(rec)
        continue
    rec['category'] = None
    rec['source'] = 'UNRESOLVED'
    still_unresolved.append(rec)
    final_records.append(rec)

with open(OUT_DIR + r'\final_summary.txt', 'w', encoding='utf-8') as f:
    f.write(f'Total: {len(final_records)}\n')
    f.write(f'By source: {Counter(r["source"].split(":")[0] for r in final_records)}\n')
    f.write(f'Still unresolved: {len(still_unresolved)}\n\n')
    for r in still_unresolved:
        f.write(f"row{r['row']} {r['date']} | {r['merchant']!r} | {r['amount']} | type={r['type']}\n")

# dump final_records for the next step (JSON transaction build) as pickle-free json
def serialize(rec):
    d = dict(rec)
    if d['date'] is not None and hasattr(d['date'], 'isoformat'):
        d['date'] = d['date'].isoformat()[:10]
    return d

with open(OUT_DIR + r'\final_records.json', 'w', encoding='utf-8') as f:
    json.dump([serialize(r) for r in final_records], f, ensure_ascii=False, indent=1)

print('done', len(final_records), 'unresolved', len(still_unresolved))
